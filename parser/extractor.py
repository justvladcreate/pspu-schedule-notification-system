# [file name]: extractor.py
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
import time
import aiofiles
import asyncio
from parser.preprocess import remove_academic_titles, clean, normalize_rooms, normalize_time

pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
pd.set_option('display.width', None)


class DataExtractor:
    def __init__(self):
        self.SCOPES = ['https://www.googleapis.com/auth/drive.readonly']
        self.file_id = "1_fsm-OxH9E9LgHnLC0iju5OlaHIv0agmM87GLvRKIAg"
        current_dir = Path(__file__).resolve().parent.parent
        self.token_file = current_dir / 'private\\token.json'
        self.credentials_file = current_dir / 'private\\credentials.json'

    async def get_service(self):
        creds = None
        try:
            creds = Credentials.from_authorized_user_file(str(self.token_file), self.SCOPES)
        except:
            pass
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(str(self.credentials_file), self.SCOPES)
                creds = flow.run_local_server(port=0)
            async with aiofiles.open(self.token_file, 'w') as token:
                await token.write(creds.to_json())
        service = build('drive', 'v3', credentials=creds, cache_discovery=False)
        return service, creds  # возвращаем и creds
    
    async def get_sheets_metadata(self):
        service, creds = await self.get_service()
        sheets_service = build('sheets', 'v4', credentials=creds, cache_discovery=False)
        
        try:
            spreadsheet = sheets_service.spreadsheets().get(
                spreadsheetId=self.file_id
            ).execute()
            
            sheets_metadata = {}
            for sheet in spreadsheet.get('sheets', []):
                sheet_props = sheet['properties']
                sheets_metadata[sheet_props['title']] = {
                    'sheetId': sheet_props['sheetId'],
                    'title': sheet_props['title'],
                    'index': sheet_props['index'],
                    'gid': sheet_props['sheetId']
                }
            
            return sheets_metadata
            
        except Exception as e:
            print(f"Ошибка при получении метаданных листов: {e}")
            return {}

    def _sync_download(self, excel_path):
        # Получаем учётные данные синхронно
        creds = None
        try:
            creds = Credentials.from_authorized_user_file(str(self.token_file), self.SCOPES)
        except:
            pass
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(str(self.credentials_file), self.SCOPES)
                creds = flow.run_local_server(port=0)
            # Сохраняем токен синхронно
            with open(self.token_file, 'w') as token:
                token.write(creds.to_json())

        # Строим сервис и скачиваем
        service = build('drive', 'v3', credentials=creds, cache_discovery=False)
        request = service.files().export_media(
            fileId=self.file_id,
            mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        with open(excel_path, 'wb') as f:
            downloader = MediaIoBaseDownload(f, request)
            done = False
            while not done:
                status, done = downloader.next_chunk()
                print(f"Download progress: {int(status.progress() * 100)}%")

    async def download_file(self, excel_path):
        await asyncio.to_thread(self._sync_download, excel_path)


    def delete_old_file(self, file_path, max_time=600):
        path = Path(file_path)

        if not path.exists():
            return

        file_time = path.stat().st_mtime
        current_time = time.time()

        if current_time - file_time > max_time:
            old_file_path = path.parent / f"old_{path.name}"

            if old_file_path.exists():
                old_file_path.unlink()

            path.rename(old_file_path)


    async def extract(self, file_path):
        sheets_metadata = await self.get_sheets_metadata()
        sheet_names = pd.ExcelFile(file_path).sheet_names
        groups_info = {}

        #Проходимся по всем листам
        for sheet_name in sheet_names[1:]:
            df = fill_merged_cells_safe(file_path, sheet_name=sheet_name)
            sheet_gid = None
            if sheet_name in sheets_metadata:
                sheet_gid = sheets_metadata[sheet_name]['gid']
            
            group_info = extraction(df, sheet_gid)
            if not group_info:
                continue
            groups_info.update(group_info)
            
        return groups_info


def clean_group_name(group_name):
    bracket_index = group_name.find(' (')
    comma_index = group_name.find(', ')+2
    if bracket_index != -1:
        return group_name[comma_index:bracket_index].strip()
    return group_name.strip()

def extraction(df, sheet_id):
    # print(df)
    # Указываем стартовые клетки
    start_cells = []
    for i, row in df.iterrows():
        # row = set(row)
        for j, cell in enumerate(row):
            if "Начало" in str(cell).strip():
                start_cells.append((i, j))
    # Указываем конечную высоту поиска
    end_row = None
    for i, row in df.iterrows():
        row = set(row)
        for j, cell in enumerate(row):
            if "Декан" in str(cell).strip():
                end_row = i
                break
        if end_row:
            break
    # Указываем конечные клетки
    end_cells = []
    for i, row in df.iterrows():
        # row = set(row)
        for j, cell in enumerate(row):
            if "форма обучения /" in str(cell).strip().lower():
                end_cells.append((end_row,j))
    # Ищем заголовки
    headers = []
    for i, row in df.iterrows():
        row_set = set()
        for j, cell in enumerate(row):
            if "семестр" in str(cell).strip().lower() and not (cell in row_set):
                headers.append((i,j))
                row_set.add(cell)

    group_info = {}

    min_count = min(len(start_cells), len(end_cells), len(headers))


    if min_count == 0:
        return group_info

    # Проходим по всем блокам по каждой клетке внутри блока
    for i in range(min_count):

        start_cell = start_cells[i]
        end_cell = end_cells[i]
        header_cell = headers[i]

        try:
            # Разбираем заголовок
            header = df.iloc[header_cell[0], header_cell[1]]

            group_name = ""
            semester = None
            additional_info = ""
            session = ""
            vacation = ""

            if "семестр" in str(header).lower():
                lines = [line.strip() for line in str(header).split("\n") if line.strip()]

                if len(lines) >= 2:
                    group_name = lines[0]
                    group_name = clean_group_name(group_name)
                    semester = lines[1]

                    try:
                        session_index = next(i for i, line in enumerate(lines) if "Экзаменационная сессия" in line)
                        vacation_index = next(i for i, line in enumerate(lines) if "Каникулы" in line)

                        additional_info = "\n".join(lines[2:session_index])
                        session = lines[session_index]
                        vacation = lines[vacation_index]
                    except StopIteration:
                        print("Не найдены сессия или каникулы в заголовке")
                else:
                    print(f"Недостаточно строк в заголовке: {len(lines)}")
        except Exception as e:
            print(f"Ошибка при обработке заголовка: {e}")
            continue

        if not group_name:
            print("Не удалось определить номер группы, пропускаем")
            continue

        try:
            subset = df.iloc[start_cell[0]:end_cell[0]+1, start_cell[1]:end_cell[1]+1]
        except Exception as e:
            print(f"Ошибка при создании subset: {e}")
            continue

        # times = []
        # subjects = []
        # teachers = []
        # rooms = []

        events = []

        # Разбираем строчки на пары
        for row in range(1, subset.shape[0]-1):
            if row >= subset.shape[0]:
                break

            abs_row = start_cell[0] + row
            # time_col = start_cell[1]
            # subject_col = start_cell[1] + 1
            # room_col = end_cell[1] - 1

            try:

                # extra
                time_val_exception = df.iloc[abs_row, start_cells[0][1]]
                room_val_exception = df.iloc[abs_row, end_cells[-1][1]]

                time_val = subset.iloc[row, 0]
                room_val = subset.iloc[row, -1]
                subject_val = subset.iloc[row, 1:-1]

                if pd.isna(time_val) or str(time_val).strip() == "":
                    continue
                if pd.isna(room_val) or str(room_val).strip() == "" or room_val is None:
                    room_val = ""

                subject_val = set(subject_val)
                subject_val = " ".join([str(x).strip() for x in subject_val if pd.notna(x)]).strip()
                if not subject_val:
                    continue

                if (time_val in subject_val) or (room_val in subject_val) and (room_val != "" and subject_val != ""):
                    if time_val in subject_val:
                        time_val = time_val_exception
                    if room_val in subject_val:
                        room_val = room_val_exception

                room_val = normalize_rooms(room_val)

                subject_val = clean(text=subject_val)
                subject_val = remove_academic_titles(text=subject_val)

                # Добавляем день недели
                day_of_week = ""
                df.iloc[:, 0] = df.iloc[:, 0].ffill()
                left_col_val = df.iloc[abs_row, 0]
                if pd.notna(left_col_val) and str(left_col_val).strip():
                    day_of_week = str(left_col_val).strip().upper()


                time_vals = time_val.split("\n")

                for time_value in time_vals:
                    time_val = normalize_time(time_value)
                    # Собираем все данные воедино
                    events.append(str("discipline info: "+day_of_week+" "+time_val+"\n"+subject_val+"\n"+"rooms info: "+", ".join(room_val)))


            except Exception as e:
                print(f"Ошибка при обработке строки {row} {group_name}: {e}")

        # print(events)
        if events:
            group_info[group_name] = {
                "events": events,
            }

    return group_info

def fill_merged_cells_safe(file_path, sheet_name):
    """
    Reads an Excel sheet and fills only the cells that belong to merged ranges.
    Genuine NaN values outside merged ranges are left untouched.
    """
    # 1. Read the sheet with pandas (merged cells appear as NaN except top-left)
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    df = df.astype(object)

    # 2. Load the same sheet with openpyxl to get merged range definitions
    wb = load_workbook(file_path, data_only=True)
    if isinstance(sheet_name, int):
        ws = wb.worksheets[sheet_name]
    else:
        ws = wb[sheet_name]

    for merged_range in ws.merged_cells.ranges:
        top_left_value = ws.cell(merged_range.min_row, merged_range.min_col).value
        start_row = merged_range.min_row - 1
        end_row = merged_range.max_row - 1
        start_col = merged_range.min_col - 1
        end_col = merged_range.max_col - 1

        if start_row < 0 or end_row >= len(df) or start_col < 0 or end_col >= len(df.columns):
            continue

        for i in range(start_row, end_row + 1):
            for j in range(start_col, end_col + 1):
                df.iat[i, j] = top_left_value
    return df